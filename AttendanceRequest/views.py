from django.shortcuts import render
from django.http import HttpResponse
import pyrebase
from CollegeAutomation.settings import firebaseConfig
import uuid
import firebase_admin
from firebase_admin import credentials
from firebase_admin import storage
import  openpyxl
from datetime import date, timedelta

firebase= pyrebase.initialize_app(firebaseConfig)
db = firebase.database()
storage  = firebase.storage()

def attendance_application(request):
    if request.method=="POST":
        Competition_Name = request.POST.get('Competition_Name')
        Organization_Name = request.POST.get('Organization_Name')
        Start_Date = request.POST.get('Start_date')
        End_date = request.POST.get('End_date')
        presence_proof = request.FILES['Attendance_Proof']
        userId = request.POST.get('localId')
        print(presence_proof)
        uid = uuid.uuid1()
        P = storage.child('Images/'+str(uid)).put(presence_proof)
        print(P)
        db.child('Attendance').child('Requests').push({
            'CompetitonName':Competition_Name,
            'Organization_Name':Organization_Name,
            'Start_Date':Start_Date,
            'End_Date':End_date,
            'presence_proof':storage.child('Images/'+str(uid)).get_url(None),
            'dean_status':0,
            'userId':userId,
            'General_Secratary_status':0  
        })
        # Fetch the service account key JSON file contents


        # Initialize the app with a service account, granting admin privileges
        
        return HttpResponse("Done")
    localId  = request.GET.get('localId')
    print(localId)
    return render(request,'Attendancerequest/request_competition_Attendance.html', {'localId':localId})

def check_attendance_hod(request):
    if request.method == "POST":
        Attendance_requests  = []
        userId  = request.POST.get('userId')
        print(userId)
        Branch_DICT = {'COMPS':6, 'IT':4, 'ETRX':3, 'EXTC':5, 'MCA':7}
        print(db.child('Users').child('Authorities').child('HOD').child(userId).get().val()['Department'])
        department  = Branch_DICT[db.child('Users').child('Authorities').child('HOD').child(userId).get().val()['Department']]
        print('Department', department)
        Attendance_request = db.child('Attendance').child('Requests').shallow().get().val()
        for i in Attendance_request:
            if db.child('Attendance').child('Requests').child(i).get().val()['dean_status'] ==  1:
                continue 
            student_id = db.child('Attendance').child('Requests').child(i).child('userId').get().val()
            if int(db.child('Users').child('Students').child(student_id).get().val()['Branch'] == int(department)):
                Student = db.child('Users').child('Students').child(student_id).get().val()
                a = db.child('Attendance').child('Requests').child(i).get().val()
                Attendance_requests.append({
                    'id':i,
                    'Student':Student,
                    'CompetitonName' : a['CompetitonName'],
                    'OrganizationName': a['Organization_Name'] ,
                    'Start_Date': a['Start_Date'],
                    'End_Date': a['End_Date'],
                    'presence_proof':a['presence_proof'],                     
                })
        return render(request, 'AttendanceRequest/check_attendance_hod.html', {'requests': Attendance_requests , 'localId':userId})    


def approve_attendance(request):
    if request.method == "POST":
        id  = request.POST.get('id')
        localId  = request.POST.get('userId')
        Attendance_requests=[]
        db.child('Attendance').child('Requests').child(id).update({'dean_status':1})
        
        print(localId)
        Branch_DICT = {'COMPS':6, 'IT':4, 'ETRX':3, 'EXTC':5, 'MCA':7}
        print(db.child('Users').child('Authorities').child('HOD').child(localId).get().val()['Department'])
        department  = Branch_DICT[db.child('Users').child('Authorities').child('HOD').child(localId).get().val()['Department']]
        print('Department', department)
        Attendance_request = db.child('Attendance').child('Requests').shallow().get().val()
        for i in Attendance_request:
            if db.child('Attendance').child('Requests').child(i).get().val()['dean_status'] ==  1:
                continue 
            student_id = db.child('Attendance').child('Requests').child(i).child('userId').get().val()
            if int(db.child('Users').child('Students').child(student_id).get().val()['Branch'] == int(department)):
                Student = db.child('Users').child('Students').child(student_id).get().val()
                a = db.child('Attendance').child('Requests').child(i).get().val()
                Attendance_requests.append({
                    'id':i,
                    'Student':Student,
                    'CompetitonName' : a['CompetitonName'],
                    'OrganizationName': a['Organization_Name'] ,
                    'Start_Date': a['Start_Date'],
                    'End_Date': a['End_Date'],
                    'presence_proof':a['presence_proof'],                     
                })
        return render(request, 'AttendanceRequest/check_attendance_hod.html', {'requests': Attendance_requests , 'localId':localId})    
        
# Create your views here.


def export_excel(request):
    if request.method == "POST":
        start_date  = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        userId  = request.POST.get('userId')
        Branch_DICT = {'COMPS':6, 'IT':4, 'ETRX':3, 'EXTC':5, 'MCA':7}
        Department  = Branch_DICT[db.child('Users').child('Authorities').child('Class_Teachers').child(userId).get().val()['Department']]
        Year  = db.child('Users').child('Authorities').child('Class_Teachers').child(userId).get().val()['Class_Teacher_Of']
        student_name = {}
        Students = db.child('Users').child('Students').shallow().get().val()
        for i in Students:
            if(db.child('Users').child('Students').child(i).get().val()['Branch'] ==Department  and db.child('Users').child('Students').child(i).get().val()['Current_Year'] == Year):
                student_name[i] = db.child('Users').child('Students').child(i).get().val()['First_Name']
        print(student_name)
        key = sorted(student_name)
        print(type(key), int(start_date.split('-')[0]), int(start_date.split('-')[1]), int(start_date.split('-')[2]))
        date1  = date(int(start_date.split('-')[0]), int(start_date.split('-')[1]), int(start_date.split('-')[2]))
        date2  = date( int(end_date.split('-')[0]), int(end_date.split('-')[1]), int(end_date.split('-')[2]))
        dates  = []       
        workbook = openpyxl.Workbook()
        worksheet = workbook.create_sheet()
        worksheet.cell(row  =   1,column =  1, value="UID")
        worksheet.cell( row = 1,column = 2 , value = "NAME") 
        for i in range(len(student_name)):
            worksheet.cell( row  = i+2,column= 1,value= key[i])
            worksheet.cell( row = i+2, column = 2 , value  = student_name[key[i]])

        for i in range(int((date2 - date1).days)):
            dates.append(str(date1+ timedelta(i)))
            worksheet.cell(row  = 1 ,column = i+3 , value  = str(date1 + timedelta(i)))
        requests  = db.child('Attendance').child('Requests').shallow().get().val()
        for i in requests:
            if db.child('Attendance').child('Requests').child(i).get().val()['userId'] in key:
                row_idx = key.index(db.child('Attendance').child('Requests').child(i).get().val()['userId'])
                req_start_date  = db.child('Attendance').child('Requests').child(i).get().val()['Start_Date']
                req_end_date  = db.child('Attendance').child('Requests').child(i).get().val()['End_Date']
                date3 = date(int(req_start_date.split('-')[0]), int(req_start_date.split('-')[1]), int(req_start_date.split('-')[2]))
                date4 = date( int(req_end_date.split('-')[0]), int(req_end_date.split('-')[1]), int(req_end_date.split('-')[2]))
                req_end_date = date4
                req_start_date = date3                
                for j in range(int((req_end_date- req_start_date).days)):
                    if(str(req_start_date+ timedelta(j))) in dates:
                        cols_idx  = dates.index(str(req_start_date+ timedelta(j)))
                        worksheet.cell(row = row_idx+2, column= cols_idx+3 ,value = "p" )

                print(req_end_date, req_start_date, start_date, end_date)
                
        workbook.save('Attendance.xlsx')
        return render( request ,'registrations/class_teacher_home.html', {'localId':userId})

